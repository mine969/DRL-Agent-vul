"""
Aggregate scan results from parallel execution into Evaluation Form.xlsx
Custom implementation with fixed report parsing
"""
import re
from pathlib import Path
from collections import defaultdict
import openpyxl

# Import functions from evaluate_fill_excel.py (resolved relative to this
# file, not the caller's cwd, so it works whether invoked as
# `python scripts/aggregate_results.py` from repo root or from inside scripts/)
import importlib.util
_EVAL_MODULE_PATH = Path(__file__).resolve().parent / "evaluate_fill_excel.py"
spec = importlib.util.spec_from_file_location("evaluate_fill_excel", str(_EVAL_MODULE_PATH))
eval_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(eval_module)

classify = eval_module.classify
impact_for = eval_module.impact_for
run_ground_truth_scan = eval_module.run_ground_truth_scan
CATEGORY_ORDER = eval_module.CATEGORY_ORDER

# Map report files to targets
REPORT_MAPPING = {
    "ecommerce": "vulnerability_report_20260810_023345.md",
    "social": "vulnerability_report_20260810_024136.md",
    "banking": "vulnerability_report_20260810_024346.md",
    "blog": "vulnerability_report_20260810_025142.md",
    "fileshare": "vulnerability_report_20260810_025430.md",
}

TARGET_NAMES = {
    "ecommerce": "E-Commerce (5002)",
    "social": "Social Media (5003)",
    "banking": "Banking (5004)",
    "blog": "Blog (5005)",
    "fileshare": "File Share (5006)",
}

APP_PATHS = {
    "ecommerce": "env/target_app_ecommerce.py",
    "social": "env/target_app_social.py",
    "banking": "env/target_app_banking.py",
    "blog": "env/target_app_blog.py",
    "fileshare": "env/target_app_fileshare.py",
}


def parse_report_findings_custom(report_filename):
    """Parse findings from a report file - handles emoji headers"""
    report_path = Path("reports") / report_filename
    
    if not report_path.exists():
        print(f"    Warning: Report not found: {report_path}")
        return []

    try:
        lines = report_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    except Exception as e:
        print(f"    Error reading report: {e}")
        return []

    findings = []
    current = {}
    in_section = False

    tech_re = re.compile(r"Technical Name\*\*: `([^`]+)`")
    url_re = re.compile(r"\*\*Vulnerable URL\*\*: `([^`]+)`")

    for line in lines:
        stripped = line.strip()
        
        # Look for "Confirmed Vulnerabilities" section (with or without emoji)
        if stripped.startswith("## ") and "Confirmed Vulnerabilities" in stripped:
            in_section = True
            continue
        
        # Exit section when we hit another ## header
        if in_section and stripped.startswith("## "):
            break
        
        if not in_section:
            continue

        # New vulnerability entry
        if stripped.startswith("### "):
            if current:
                findings.append(current)
            current = {}
            continue

        # Extract technical name
        tech_match = tech_re.search(stripped)
        if tech_match:
            current["technical_name"] = tech_match.group(1)
            continue

        # Extract URL
        url_match = url_re.search(stripped)
        if url_match:
            current["url"] = url_match.group(1)
            continue

    # Don't forget the last one
    if current:
        findings.append(current)

    return [f for f in findings if f.get("technical_name")]


# Ground truth from application files
print("Parsing ground truth from application files...")
GROUND_TRUTH = {}
for key, app_path in APP_PATHS.items():
    print(f"  Scanning {TARGET_NAMES[key]}...")
    GROUND_TRUTH[key] = run_ground_truth_scan(app_path)
    total = sum(len(v) for v in GROUND_TRUTH[key].values())
    print(f"    Found {total} vulnerabilities")

# Parse detected vulnerabilities from reports
print("\nParsing detected vulnerabilities from reports...")
DETECTED = {}
for key, report_file in REPORT_MAPPING.items():
    print(f"  Parsing {TARGET_NAMES[key]} report...")
    findings = parse_report_findings_custom(report_file)
    
    detected_by_category = defaultdict(set)
    
    for finding in findings:
        technical = finding.get("technical_name")
        if not technical:
            continue
        category = classify(technical)
        detected_by_category[category].add(technical)
    
    DETECTED[key] = detected_by_category
    total = sum(len(v) for v in detected_by_category.values())
    print(f"    Found {total} detected vulnerabilities")

# Generate Excel rows
print("\nGenerating Excel rows...")
rows = []
site_index = 1

for key in ["ecommerce", "social", "banking", "blog", "fileshare"]:
    gt = GROUND_TRUTH.get(key, {})
    det = DETECTED.get(key, {})
    
    categories = [c for c in CATEGORY_ORDER if c in gt]
    extra = sorted([c for c in gt.keys() if c not in CATEGORY_ORDER])
    categories += extra
    
    first_row = True
    for category in categories:
        total = len(gt.get(category, set()))
        if total == 0:
            continue
        detected_set = det.get(category, set())
        detected_count = min(len(detected_set), total)
        percent = (detected_count / total) * 100 if total else 0.0
        impact = impact_for(category)
        
        rows.append([
            site_index if first_row else None,
            TARGET_NAMES[key] if first_row else None,
            category,
            total,
            detected_count,
            f"{percent:.0f}%",
            impact.title() if isinstance(impact, str) else impact,
        ])
        first_row = False
    
    site_index += 1

# Update Excel file
print(f"\nUpdating Evaluation Form.xlsx with {len(rows)} rows...")
wb = openpyxl.load_workbook("Evaluation Form.xlsx")
ws = wb.active
ws.delete_rows(2, ws.max_row)

for idx, row in enumerate(rows, start=2):
    for col, value in enumerate(row, start=1):
        ws.cell(row=idx, column=col, value=value)

wb.save("Evaluation Form.xlsx")

print(f"✅ Successfully populated Evaluation Form.xlsx")
print(f"\n📊 Final Summary:")
for key in ["ecommerce", "social", "banking", "blog", "fileshare"]:
    gt = GROUND_TRUTH.get(key, {})
    det = DETECTED.get(key, {})
    total_vulns = sum(len(v) for v in gt.values())
    detected_vulns = sum(len(v) for v in det.values())
    detection_rate = (detected_vulns / total_vulns * 100) if total_vulns > 0 else 0
    print(f"  {TARGET_NAMES[key]}: {detected_vulns}/{total_vulns} ({detection_rate:.1f}%)")
