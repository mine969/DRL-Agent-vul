import argparse
import io
import os
import re
import sys
import time
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
import requests
import torch

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from agent.improved_dqn_agent import ImprovedDQNAgent
from env.web_sec_env import WebSecEnv
from utils.model_loader import load_model_smart


@dataclass
class Target:
    name: str
    url: str
    app: str
    port: int


TARGETS = [
    Target("E-Commerce", "http://localhost:5002", "env/target_app_ecommerce.py", 5002),
    Target("Social Media", "http://localhost:5003", "env/target_app_social.py", 5003),
    Target("Banking", "http://localhost:5004", "env/target_app_banking.py", 5004),
    Target("Blog", "http://localhost:5005", "env/target_app_blog.py", 5005),
    Target("File Share", "http://localhost:5006", "env/target_app_fileshare.py", 5006),
]

DEFAULT_MODEL_PATH = "checkpoints/improved_mock_ep10000.pth"
DEFAULT_WORKBOOK_PATH = "research/results/Evaluation Form.xlsx"
DEFAULT_RUNS = 5
DEFAULT_SCAN_STEPS = 15
FULL_POTENTIAL_SCAN_STEPS = 60

TYPE_ORDER = [
    "SQL Injection",
    "Cross-Site Scripting (XSS)",
    "Insecure Direct Object Reference (IDOR)",
    "Broken Access Control (BAC)",
    "Mass Assignment",
    "Business Logic",
    "Race Condition",
    "Cross-Site Request Forgery (CSRF)",
    "Path Traversal",
    "File Upload",
    "Server-Side Template Injection (SSTI)",
    "Server-Side Request Forgery (SSRF)",
    "Command Injection",
    "Insecure Deserialization",
    "Weak Password",
    "Session Fixation",
    "Weak Reset Token",
    "OAuth Bypass",
    "JWT Bypass",
    "SAML Bypass",
    "Information Disclosure",
    "Sensitive Data Exposure",
    "Auth Bypass",
]

IMPACT_MAP = {
    "SQL Injection": "Critical",
    "Cross-Site Scripting (XSS)": "High",
    "Insecure Direct Object Reference (IDOR)": "Medium",
    "Broken Access Control (BAC)": "High",
    "Mass Assignment": "High",
    "Business Logic": "Medium",
    "Race Condition": "Medium",
    "Cross-Site Request Forgery (CSRF)": "Medium",
    "Path Traversal": "High",
    "File Upload": "Critical",
    "Server-Side Template Injection (SSTI)": "Critical",
    "Server-Side Request Forgery (SSRF)": "High",
    "Command Injection": "Critical",
    "Insecure Deserialization": "Critical",
    "Weak Password": "High",
    "Session Fixation": "High",
    "Weak Reset Token": "High",
    "OAuth Bypass": "High",
    "JWT Bypass": "High",
    "SAML Bypass": "High",
    "Information Disclosure": "High",
    "Sensitive Data Exposure": "High",
    "Auth Bypass": "High",
}


def classify_marker(text: str) -> Optional[str]:
    if not text:
        return None
    t = text.lower()
    if "sqli" in t or "sql injection" in t or "sql" in t:
        return "SQL Injection"
    if "xss" in t:
        return "Cross-Site Scripting (XSS)"
    if "idor" in t or "insecure direct object" in t:
        return "Insecure Direct Object Reference (IDOR)"
    if "csrf" in t:
        return "Cross-Site Request Forgery (CSRF)"
    if "upload" in t:
        return "File Upload"
    if "traversal" in t or "path traversal" in t:
        return "Path Traversal"
    if "ssti" in t or "template injection" in t:
        return "Server-Side Template Injection (SSTI)"
    if "ssrf" in t or "server-side request forgery" in t:
        return "Server-Side Request Forgery (SSRF)"
    if "command injection" in t or "cmd_injection" in t or "cmd injection" in t:
        return "Command Injection"
    if "deserialization" in t:
        return "Insecure Deserialization"
    if "mass assignment" in t:
        return "Mass Assignment"
    if "race condition" in t or "race" in t:
        return "Race Condition"
    if (
        "business logic" in t
        or "price manipulation" in t
        or "payment bypass" in t
        or "negative quantity" in t
        or "coupon" in t
    ):
        return "Business Logic"
    if "weak password" in t or "weak auth" in t:
        return "Weak Password"
    if "session fixation" in t:
        return "Session Fixation"
    if "reset token" in t or "password reset" in t or "predictable reset" in t:
        return "Weak Reset Token"
    if "oauth" in t or "state" in t and "oauth" in t:
        return "OAuth Bypass"
    if "jwt" in t:
        return "JWT Bypass"
    if "saml" in t:
        return "SAML Bypass"
    if "info disclosure" in t or "information disclosure" in t or "secret" in t:
        return "Information Disclosure"
    if "sensitive data" in t or "api key" in t:
        return "Sensitive Data Exposure"
    if "broken access control" in t or "bac" in t or "admin users" in t:
        return "Broken Access Control (BAC)"
    if "auth bypass" in t:
        return "Auth Bypass"
    return None


def extract_route_blocks(file_path: str) -> List[str]:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.read().splitlines()

    blocks: List[str] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.lstrip().startswith("@app.route"):
            j = i + 1
            while j < len(lines) and not lines[j].lstrip().startswith("def "):
                j += 1
            if j >= len(lines):
                break
            func_indent = len(lines[j]) - len(lines[j].lstrip())
            k = j + 1
            while k < len(lines):
                if lines[k].lstrip().startswith("@app.route") and (
                    len(lines[k]) - len(lines[k].lstrip()) == func_indent
                ):
                    break
                k += 1
            blocks.append("\n".join(lines[j:k]))
            i = k
        else:
            i += 1
    return blocks


def extract_markers(block: str) -> Set[str]:
    markers: Set[str] = set()
    markers.update(re.findall(r'X-Vuln-Confirmed"\]\s*=\s*"([^"]+)"', block))
    markers.update(re.findall(r"X-Vuln-Confirmed'\]\s*=\s*'([^']+)'", block))
    markers.update(re.findall(r"CTF\{[^}]+\}", block))
    markers.update(re.findall(r'"vuln"\s*:\s*"([^"]+)"', block))
    markers.update(re.findall(r"'vuln'\s*:\s*'([^']+)'", block))

    for line in block.splitlines():
        if "VULN" in line or "VULNERABILITY" in line:
            markers.add(line.strip())
    return markers


def ground_truth_counts(file_path: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    blocks = extract_route_blocks(file_path)
    for block in blocks:
        types: Set[str] = set()
        for marker in extract_markers(block):
            vuln_type = classify_marker(marker)
            if vuln_type:
                types.add(vuln_type)
        for vuln_type in types:
            counts[vuln_type] = counts.get(vuln_type, 0) + 1
    return counts


def is_reachable(url: str, timeout: float = 1.0) -> bool:
    try:
        requests.get(url, timeout=timeout)
        return True
    except requests.RequestException:
        return False


def wait_for_target(url: str, timeout_seconds: float = 6.0) -> bool:
    start = time.time()
    while time.time() - start < timeout_seconds:
        if is_reachable(url):
            return True
        time.sleep(0.5)
    return False


def start_target(target: Target) -> Tuple[Optional[subprocess.Popen], bool, str]:
    if is_reachable(target.url):
        return None, True, "already_running"
    process = subprocess.Popen(
        [sys.executable, target.app],
        cwd=os.getcwd(),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    ok = wait_for_target(target.url)
    return process, ok, "started" if ok else "failed"


def stop_process(process: Optional[subprocess.Popen]) -> None:
    if not process:
        return
    try:
        process.terminate()
        process.wait(timeout=3)
    except Exception:
        try:
            process.kill()
            process.wait(timeout=3)
        except Exception:
            pass


def scan_model_counts(
    target: Target, model_path: str, scan_steps: int, full_potential: bool
) -> Dict[str, int]:
    env = WebSecEnv(target_url=target.url, mode="mock_targets")
    env.max_steps_per_episode = max(scan_steps, getattr(env, "max_steps_per_episode", scan_steps))

    agent = ImprovedDQNAgent(state_dim=15, action_dim=50)
    checkpoint = torch.load(model_path, map_location=agent.device)
    if isinstance(checkpoint, dict) and "q_network_state_dict" in checkpoint:
        agent.q_network.load_state_dict(checkpoint["q_network_state_dict"])
        if "target_network_state_dict" in checkpoint:
            agent.target_network.load_state_dict(checkpoint["target_network_state_dict"])
        else:
            agent.target_network.load_state_dict(agent.q_network.state_dict())
        print(f"Loaded exact checkpoint: {model_path}")
    else:
        load_model_smart(agent, model_path=model_path, auto_checkpoint=False, verbose=True)

    network = getattr(agent, "q_network", None)
    if network:
        if full_potential:
            network.train()
        else:
            network.eval()

    try:
        env.action_login_valid()
    except Exception:
        pass

    buckets: Dict[str, Set[str]] = {}
    state, _ = env.reset()
    for _ in range(scan_steps):
        action = agent.act(state, training=full_potential)
        next_state, _, terminated, truncated, info = env.step(action)

        response = getattr(env, "last_response", None)
        vuln_id = None

        if response is not None and response.headers.get("X-Vuln-Confirmed"):
            vuln_id = response.headers.get("X-Vuln-Confirmed")

        flags = None
        if isinstance(info, dict):
            flags = info.get("flags")

        if not vuln_id and flags:
            vuln_id = flags[0]

        if not vuln_id and response is not None and response.text:
            match = re.search(r"CTF\{[^}]+\}", response.text)
            if match:
                vuln_id = match.group(0)

        if vuln_id:
            real_action_id = action
            if hasattr(env, "mock_action_map"):
                real_action_id = env.mock_action_map.get(action, action)
            action_func = env.action_book.get(real_action_id)
            action_name = action_func.__name__ if action_func else str(action)
            vuln_type = classify_marker(action_name) or classify_marker(vuln_id)
            if vuln_type:
                key = f"{vuln_type}|{action_name}|{vuln_id}"
                buckets.setdefault(vuln_type, set()).add(key)

        if terminated or truncated:
            try:
                env.action_login_valid()
            except Exception:
                pass
            state, _ = env.reset()
            continue

        state = next_state

    env.close()
    return {k: len(v) for k, v in buckets.items()}


def sort_types(types: List[str]) -> List[str]:
    index = {name: i for i, name in enumerate(TYPE_ORDER)}
    return sorted(types, key=lambda name: index.get(name, len(TYPE_ORDER)))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run repeated model evaluations and write average findings to the workbook."
    )
    parser.add_argument("--model", default=DEFAULT_MODEL_PATH, help="Checkpoint to evaluate")
    parser.add_argument("--runs", type=int, default=DEFAULT_RUNS, help="Repeated runs per target")
    parser.add_argument(
        "--steps",
        type=int,
        default=DEFAULT_SCAN_STEPS,
        help="Total agent action budget per run",
    )
    parser.add_argument(
        "--full-potential",
        action="store_true",
        help="Use a larger step budget and noisy-network exploration during evaluation",
    )
    parser.add_argument(
        "--workbook",
        default=DEFAULT_WORKBOOK_PATH,
        help="Workbook to update with averaged results",
    )
    return parser.parse_args()


def _format_average(value: float) -> float | int:
    rounded = round(value, 1)
    if abs(rounded - round(rounded)) < 1e-9:
        return int(round(rounded))
    return rounded


def main() -> int:
    args = parse_args()
    model_path = str(Path(args.model))
    workbook_path = Path(args.workbook)
    runs = max(1, int(args.runs))
    full_potential = bool(args.full_potential)
    scan_steps = max(1, int(args.steps))
    if full_potential and scan_steps == DEFAULT_SCAN_STEPS:
        scan_steps = FULL_POTENTIAL_SCAN_STEPS

    if not Path(model_path).exists():
        raise FileNotFoundError(f"Model checkpoint not found: {model_path}")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    print(f"Using model: {model_path}")
    print(f"Repeated runs per target: {runs}")
    print(f"Scan steps per run: {scan_steps}")
    print(f"Full potential mode: {full_potential}")

    processes: List[Optional[subprocess.Popen]] = []
    try:
        for target in TARGETS:
            proc, ok, status = start_target(target)
            processes.append(proc)
            if not ok:
                print(f"Failed to start {target.name}: {status}")
                return 1

        ground_truth: Dict[str, Dict[str, int]] = {}
        model_runs: Dict[str, List[Dict[str, int]]] = {}

        for target in TARGETS:
            ground_truth[target.name] = ground_truth_counts(target.app)

        for target in TARGETS:
            model_runs[target.name] = []
            for run_index in range(runs):
                print(f"[{target.name}] Run {run_index + 1}/{runs}")
                model_runs[target.name].append(
                    scan_model_counts(target, model_path, scan_steps, full_potential)
                )

        wb = openpyxl.load_workbook(workbook_path)
        ws = wb.active

        for r in range(2, ws.max_row + 1):
            for c in range(1, 8):
                ws.cell(r, c).value = None

        row = 2
        site_index = 1
        for target in TARGETS:
            gt = ground_truth.get(target.name, {})
            types = sort_types(list(gt.keys()))
            for idx, vuln_type in enumerate(types):
                total = gt.get(vuln_type, 0)
                per_run_detected = [
                    min(run.get(vuln_type, 0), total)
                    for run in model_runs.get(target.name, [])
                ]
                avg_detected = (
                    sum(per_run_detected) / len(per_run_detected)
                    if per_run_detected
                    else 0.0
                )
                percent = 0.0 if total == 0 else (avg_detected / total) * 100.0

                if idx == 0:
                    ws.cell(row, 1).value = site_index
                    ws.cell(row, 2).value = f"{target.name} ({target.port})"
                else:
                    ws.cell(row, 1).value = None
                    ws.cell(row, 2).value = None

                ws.cell(row, 3).value = vuln_type
                ws.cell(row, 4).value = total
                ws.cell(row, 5).value = _format_average(avg_detected)
                ws.cell(row, 6).value = f"{percent:.1f}%"
                ws.cell(row, 7).value = IMPACT_MAP.get(vuln_type, "Medium")
                row += 1
            site_index += 1

        wb.save(workbook_path)
        print(f"Updated workbook: {workbook_path}")
        return 0
    finally:
        for proc in processes:
            stop_process(proc)


if __name__ == "__main__":
    raise SystemExit(main())
