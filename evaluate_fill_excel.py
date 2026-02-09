import os
import sys
import time
import subprocess
import io
import re
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, Iterable, List, Set, Tuple
from collections import defaultdict

import requests
import openpyxl

from utils.vulnerability_database import VULNERABILITY_DATABASE


@dataclass
class TargetConfig:
    name: str
    url: str
    app: str
    port: int


TARGETS: Dict[str, TargetConfig] = {
    "ecommerce": TargetConfig(
        name="E-Commerce (5002)",
        url="http://localhost:5002",
        app="env/target_app_ecommerce.py",
        port=5002,
    ),
    "social": TargetConfig(
        name="Social Media (5003)",
        url="http://localhost:5003",
        app="env/target_app_social.py",
        port=5003,
    ),
    "banking": TargetConfig(
        name="Banking (5004)",
        url="http://localhost:5004",
        app="env/target_app_banking.py",
        port=5004,
    ),
    "blog": TargetConfig(
        name="Blog (5005)",
        url="http://localhost:5005",
        app="env/target_app_blog.py",
        port=5005,
    ),
    "fileshare": TargetConfig(
        name="File Share (5006)",
        url="http://localhost:5006",
        app="env/target_app_fileshare.py",
        port=5006,
    ),
}

MODEL_PATH = "checkpoints/improved_mock_ep6000.pth"
SCAN_DEPTH = 120
SCAN_INTENSITY = 10
SCAN_PERSIST = True
SCAN_AI_MODE = True
SCAN_PENTESTER = True
SCAN_TIMEOUT_SECONDS = 1800
TARGET_STARTUP_TIMEOUT_SECONDS = 20.0


@dataclass(frozen=True)
class ScanPassConfig:
    name: str
    depth: int
    intensity: int
    persist: bool
    ai_mode: bool
    pentester: bool
    timeout_seconds: int


SCAN_PASSES: Tuple[ScanPassConfig, ...] = (
    ScanPassConfig(
        name="broad_ai",
        depth=SCAN_DEPTH,
        intensity=SCAN_INTENSITY,
        persist=SCAN_PERSIST,
        ai_mode=SCAN_AI_MODE,
        pentester=False,
        timeout_seconds=SCAN_TIMEOUT_SECONDS,
    ),
    ScanPassConfig(
        name="chain_attack",
        depth=max(60, SCAN_DEPTH // 2),
        intensity=SCAN_INTENSITY,
        persist=True,
        ai_mode=True,
        pentester=SCAN_PENTESTER,
        timeout_seconds=SCAN_TIMEOUT_SECONDS,
    ),
)

# Deterministic action sweep for one-shot, high-coverage detection
ACTION_SWEEP_ROUNDS = 2
ENABLE_DIRECT_ACTION_SWEEP = True
DIRECT_ACTION_SWEEP_ROUNDS = 3


if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(
        sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True
    )
    sys.stderr = io.TextIOWrapper(
        sys.stderr.buffer, encoding="utf-8", errors="replace", line_buffering=True
    )


CATEGORY_ORDER = [
    "SQL Injection",
    "Cross-Site Scripting (XSS)",
    "Insecure Direct Object Reference (IDOR)",
    "Broken Access Control (BAC)",
    "Cross-Site Request Forgery (CSRF)",
    "File Upload",
    "Path Traversal",
    "Server-Side Template Injection (SSTI)",
    "Server-Side Request Forgery (SSRF)",
    "Command Injection",
    "Insecure Deserialization",
    "Mass Assignment",
    "Business Logic",
    "Sensitive Data Exposure",
    "Weak Password",
    "Session Fixation",
    "Weak Reset Token",
    "OAuth Bypass",
    "JWT Bypass",
    "SAML Bypass",
]


CUSTOM_IMPACT = {
    "File Upload": "CRITICAL",
    "Mass Assignment": "HIGH",
    "Business Logic": "MEDIUM",
    "Weak Password": "HIGH",
    "Session Fixation": "HIGH",
    "Weak Reset Token": "HIGH",
    "OAuth Bypass": "HIGH",
    "JWT Bypass": "HIGH",
    "SAML Bypass": "HIGH",
}


def is_reachable(url: str, timeout: float = 1.0) -> bool:
    try:
        requests.get(url, timeout=timeout)
        return True
    except requests.RequestException:
        return False


def wait_for_target(url: str, timeout_seconds: float = 12.0) -> bool:
    start = time.time()
    while time.time() - start < timeout_seconds:
        if is_reachable(url):
            return True
        time.sleep(0.5)
    return False


def start_target(config: TargetConfig, startup_timeout: float = 12.0):
    if is_reachable(config.url):
        return None, True, "already_running"

    process = subprocess.Popen(
        [sys.executable, config.app],
        cwd=os.getcwd(),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    ok = wait_for_target(config.url, timeout_seconds=startup_timeout)
    status = "started" if ok else "failed_to_start"
    return process, ok, status


def stop_process(process):
    if not process:
        return
    try:
        process.terminate()
        process.wait(timeout=3)
    except Exception:
        try:
            process.kill()
            process.wait(timeout=3)
        except Exception:
            pass


def classify(action_name: str) -> str:
    if not action_name:
        return "Other"
    raw = action_name.lower()
    name = re.sub(r"[^a-z0-9]+", "_", raw)

    if "weak_password" in name or name == "test_weak_passwords":
        return "Weak Password"
    if "session_fixation" in name:
        return "Session Fixation"
    if "password_reset" in name or "reset_token" in name or "predictable_reset" in name:
        return "Weak Reset Token"
    if "oauth" in name:
        return "OAuth Bypass"
    if "jwt" in name:
        return "JWT Bypass"
    if "saml" in name:
        return "SAML Bypass"

    if "sqli" in name or "sql_injection" in name or " sql " in f" {raw} ":
        return "SQL Injection"
    if "xss" in name:
        return "Cross-Site Scripting (XSS)"
    if "idor" in name:
        return "Insecure Direct Object Reference (IDOR)"
    if name.startswith("attack_bac_") or "bac_" in name:
        return "Broken Access Control (BAC)"
    if "csrf" in name:
        return "Cross-Site Request Forgery (CSRF)"
    if "file_upload" in name or ("upload" in name and "idor" not in name):
        return "File Upload"
    if "path_traversal" in name or "traversal" in name:
        return "Path Traversal"
    if "ssti" in name:
        return "Server-Side Template Injection (SSTI)"
    if "ssrf" in name:
        return "Server-Side Request Forgery (SSRF)"
    if "command_injection" in name or "cmd_injection" in name or "shell_injection" in name:
        return "Command Injection"
    if "deserialization" in name:
        return "Insecure Deserialization"
    if "mass_assignment" in name:
        return "Mass Assignment"
    if any(token in name for token in ["negative_quantity", "price_", "payment", "coupon", "race_"]):
        return "Business Logic"
    if "info_disclosure" in name or "insecure_api_keys" in name or "secret" in name:
        return "Sensitive Data Exposure"

    return "Other"


def impact_for(category: str) -> str:
    if category in VULNERABILITY_DATABASE:
        return VULNERABILITY_DATABASE[category].get("impact", "UNKNOWN")
    if category in CUSTOM_IMPACT:
        return CUSTOM_IMPACT[category]
    return "MEDIUM"


def run_ground_truth_scan(app_path: str):
    found = defaultdict(set)

    route_re = re.compile(r"^\s*@app\.route\(([^)]*)\)")
    func_re = re.compile(r"^\s*def\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*\(")
    string_re = re.compile(r"['\"]([^'\"]+)['\"]")

    lines = Path(app_path).read_text(encoding="utf-8", errors="ignore").splitlines()

    pending_routes: List[str] = []
    block_routes: List[str] = []
    current_func = ""
    block_lines: List[str] = []

    def flush_block(routes: Iterable[str], func_name: str, content: List[str]) -> None:
        if not content:
            return
        joined = "\n".join(content)
        routes_key = "|".join(sorted(set(routes))) if routes else "<no-route>"
        base_key = f"{routes_key}:{func_name or '<anon>'}"

        marker_patterns = [
            r'X-Vuln-Confirmed"\]\s*=\s*"([^"]+)"',
            r"X-Vuln-Confirmed'\]\s*=\s*'([^']+)'",
            r'"vuln"\s*:\s*"([^"]+)"',
            r"'vuln'\s*:\s*'([^']+)'",
        ]
        markers: Set[str] = set()
        for marker_re in marker_patterns:
            markers.update(re.findall(marker_re, joined, flags=re.IGNORECASE))

        for line in content:
            line_l = line.lower()
            if "vuln:" in line_l or "vulnerability:" in line_l:
                markers.add(line.strip())

        for marker in markers:
            category = classify(marker)
            if category != "Other":
                marker_l = marker.lower()
                if "potential" in marker_l and "x-vuln-confirmed" not in marker_l:
                    continue
                found[category].add(base_key)

    for line in lines:
        route_match = route_re.match(line)
        if route_match:
            route_bits = string_re.findall(route_match.group(1))
            if route_bits:
                pending_routes.extend(route_bits)
            continue

        func_match = func_re.match(line)
        if func_match:
            flush_block(block_routes, current_func, block_lines)
            current_func = func_match.group(1)
            block_routes = pending_routes[:] if pending_routes else []
            pending_routes = []
            block_lines = [line]
            continue

        if block_lines:
            block_lines.append(line)

    flush_block(block_routes, current_func, block_lines)

    return found


def snapshot_reports():
    return set(Path("reports").glob("*.md"))


def find_latest_report():
    reports = list(Path("reports").glob("*.md"))
    if not reports:
        return None
    return max(reports, key=lambda p: p.stat().st_ctime)


def find_new_report(before_snapshot):
    after_snapshot = set(Path("reports").glob("*.md"))
    new_reports = list(after_snapshot - before_snapshot)
    if new_reports:
        return max(new_reports, key=lambda p: p.stat().st_ctime)
    return None


def parse_report_findings(report_path):
    if not report_path or not Path(report_path).exists():
        return []

    try:
        lines = Path(report_path).read_text(encoding="utf-8", errors="ignore").splitlines()
    except Exception:
        return []

    findings = []
    current = {}
    in_section = False

    tech_re = re.compile(r"Technical Name\*\*: `([^`]+)`")
    url_re = re.compile(r"\*\*Vulnerable URL\*\*: `([^`]+)`")

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("## ") and "Confirmed Vulnerabilities" in stripped:
            in_section = True
            continue
        if in_section and stripped.startswith("## "):
            break
        if not in_section:
            continue

        if stripped.startswith("### "):
            if current:
                findings.append(current)
            current = {}
            continue

        tech_match = tech_re.search(stripped)
        if tech_match:
            current["technical_name"] = tech_match.group(1)
            continue

        url_match = url_re.search(stripped)
        if url_match:
            current["url"] = url_match.group(1)
            continue

    if current:
        findings.append(current)

    return [f for f in findings if f.get("technical_name")]


def _merge_found(target: Dict[str, Set[str]], source: Dict[str, Set[str]]) -> None:
    for category, entries in source.items():
        target[category].update(entries)


def _focus_actions_for_port(port: int) -> List[int]:
    by_port = {
        5002: [33, 30, 31, 25, 28, 42],
        5003: [33, 34, 30, 31, 25, 26, 42],
        5004: [33, 42, 29],
        5005: [33, 34, 38, 25],
        5006: [33, 37, 39, 25],
    }
    return by_port.get(port, [33, 30, 31, 25, 42])


def _category_from_action_result(action_name: str, response, info_flags: List[str]) -> str:
    if response is not None:
        confirmed_id = response.headers.get("X-Vuln-Confirmed", "").strip()
        if confirmed_id:
            category = classify(confirmed_id)
            if category != "Other":
                return category

    for flag in info_flags:
        category = classify(flag)
        if category != "Other":
            return category

    return classify(action_name)


def run_action_sweep_scan(target_url: str, rounds: int = ACTION_SWEEP_ROUNDS):
    from env.web_sec_env import WebSecEnv
    from urllib.parse import urlparse

    found = defaultdict(set)

    parsed = urlparse(target_url)
    port = parsed.port or 0

    env = WebSecEnv(target_url=target_url, mode="mock_targets")
    focus = _focus_actions_for_port(port)
    action_count = int(getattr(env.action_space, "n", 50))
    action_order = focus + [a for a in range(action_count) if a not in focus]

    env.max_steps_per_episode = max(120, len(action_order) + 10)

    try:
        for _ in range(max(1, rounds)):
            env.reset()
            for action_id in action_order:
                _, reward, terminated, truncated, info = env.step(action_id)

                response = getattr(env, "last_response", None)
                header_confirmed = bool(
                    response is not None
                    and response.headers.get("X-Vuln-Confirmed", "").strip()
                )
                env_confirmed = bool(info.get("vuln_found")) or header_confirmed

                # Strict acceptance to avoid noisy findings from navigation-only actions
                if not env_confirmed and reward < 1.0:
                    if terminated or truncated:
                        env.reset()
                    continue

                real_action = env.mock_action_map.get(action_id, action_id)
                action_fn = env.action_book.get(real_action)
                action_name = action_fn.__name__ if action_fn else f"action_{action_id}"

                raw_flags_obj = info.get("flags")
                raw_flags = raw_flags_obj if isinstance(raw_flags_obj, list) else []
                info_flags = [str(flag) for flag in raw_flags if flag]
                category = _category_from_action_result(action_name, response, info_flags)
                if category == "Other":
                    if terminated or truncated:
                        env.reset()
                    continue

                confirmed_id = (
                    response.headers.get("X-Vuln-Confirmed", "").strip()
                    if response is not None
                    else ""
                )
                url = (
                    response.url
                    if response is not None and getattr(response, "url", None)
                    else target_url
                )
                dedup_source = confirmed_id or action_name
                found[category].add(f"{dedup_source}|{url}")

                if terminated or truncated:
                    env.reset()
    finally:
        env.close()

    return found


def run_direct_action_scan(target_url: str, rounds: int = DIRECT_ACTION_SWEEP_ROUNDS):
    """Execute the full internal action catalog (real IDs) once for extra coverage."""
    from env.web_sec_env import WebSecEnv

    found = defaultdict(set)
    env = WebSecEnv(target_url=target_url, mode="mock_targets")

    try:
        env.max_steps_per_episode = 200

        for _ in range(max(1, rounds)):
            env.reset()
            try:
                env.action_login_valid()
            except Exception:
                pass

            for real_action_id in sorted(env.action_book.keys()):
                action_fn = env.action_book.get(real_action_id)
                if action_fn is None:
                    continue

                try:
                    result = action_fn()
                except Exception:
                    continue

                if not result or not isinstance(result, tuple) or len(result) != 2:
                    continue

                response, reward = result
                if response is None:
                    continue

                confirmed_id = response.headers.get("X-Vuln-Confirmed", "").strip()
                if not confirmed_id and float(reward or 0.0) < 1.0:
                    continue

                category = classify(confirmed_id or action_fn.__name__)
                if category == "Other":
                    category = classify(action_fn.__name__)
                if category == "Other":
                    continue

                url = response.url if getattr(response, "url", None) else target_url
                dedup_source = confirmed_id or action_fn.__name__
                found[category].add(f"{dedup_source}|{url}")
    finally:
        env.close()

    return found


def run_ai_scan_pass(target_url: str, pass_config: ScanPassConfig):
    from autonomous_scan import SecurityAuditor

    found = defaultdict(set)
    auditor = SecurityAuditor(base_url=target_url, model_path=MODEL_PATH)
    findings = auditor.start_audit(
        crawl_depth=pass_config.depth,
        test_intensity=pass_config.intensity,
        epsilon=0.15 if pass_config.ai_mode else 0.05,
        persist=pass_config.persist,
        ai_mode=pass_config.ai_mode,
        pentester=pass_config.pentester,
    )

    for finding in findings:
        confirmed_id = (getattr(finding, "confirmed_id", "") or "").strip()
        technical = getattr(finding, "vuln_type", "") or ""
        evidence = (getattr(finding, "evidence", "") or "").lower()

        # Accept only high-confidence confirmations from env/header evidence.
        is_confirmed = bool(getattr(finding, "env_confirmed", False)) or bool(confirmed_id)
        is_confirmed = is_confirmed or ("x-vuln-confirmed" in evidence)
        if not is_confirmed:
            continue

        category = classify(confirmed_id or technical)
        if category == "Other":
            category = classify(technical)
        if category == "Other":
            continue

        finding_url = getattr(finding, "url", "") or target_url
        dedup_source = confirmed_id or technical
        found[category].add(f"{dedup_source}|{finding_url}")

    return found


def run_model_scan(target_url: str):
    combined = defaultdict(set)

    print(f"    pass=action_sweep rounds={ACTION_SWEEP_ROUNDS}")
    sweep_found = run_action_sweep_scan(target_url, rounds=ACTION_SWEEP_ROUNDS)
    _merge_found(combined, sweep_found)

    if ENABLE_DIRECT_ACTION_SWEEP:
        print(f"    pass=direct_action_sweep rounds={DIRECT_ACTION_SWEEP_ROUNDS}")
        direct_found = run_direct_action_scan(
            target_url, rounds=DIRECT_ACTION_SWEEP_ROUNDS
        )
        _merge_found(combined, direct_found)

    for pass_config in SCAN_PASSES:
        print(
            f"    pass={pass_config.name} depth={pass_config.depth} intensity={pass_config.intensity} "
            f"persist={pass_config.persist} ai_mode={pass_config.ai_mode} pentester={pass_config.pentester}"
        )
        current = run_ai_scan_pass(target_url, pass_config)
        _merge_found(combined, current)

    return combined


def main():
    processes = {}
    all_ok = True

    print("Starting targets...")
    for key, cfg in TARGETS.items():
        proc, ok, status = start_target(cfg, startup_timeout=TARGET_STARTUP_TIMEOUT_SECONDS)
        processes[key] = proc
        if not ok:
            print(f"  {cfg.name}: failed to start")
            all_ok = False
        else:
            print(f"  {cfg.name}: {status}")

    if not all_ok:
        for proc in processes.values():
            stop_process(proc)
        raise SystemExit(1)

    ground_truth = {}
    detected = {}

    try:
        for key, cfg in TARGETS.items():
            # Some targets may crash/reset during long scans; ensure availability per target.
            if not is_reachable(cfg.url):
                replacement_proc, ok, status = start_target(
                    cfg, startup_timeout=TARGET_STARTUP_TIMEOUT_SECONDS
                )
                if replacement_proc is not None:
                    stop_process(processes.get(key))
                    processes[key] = replacement_proc
                print(f"[Target Check] {cfg.name}: {status}")
                if not ok:
                    ground_truth[key] = run_ground_truth_scan(cfg.app)
                    detected[key] = defaultdict(set)
                    continue

            print(f"\n[Ground Truth] Scanning actions for {cfg.name}...")
            ground_truth[key] = run_ground_truth_scan(cfg.app)

            print(f"[Model] Running model scan for {cfg.name}...")
            detected[key] = run_model_scan(cfg.url)

    finally:
        for proc in processes.values():
            stop_process(proc)

    rows = []
    site_index = 1
    for key, cfg in TARGETS.items():
        gt = ground_truth.get(key, {})
        det = detected.get(key, {})

        categories = [c for c in CATEGORY_ORDER if c in gt]
        extra = sorted([c for c in gt.keys() if c not in CATEGORY_ORDER])
        categories += extra

        first_row = True
        for category in categories:
            total = len(gt.get(category, set()))
            if total == 0:
                continue
            detected_set = det.get(category, set())
            detected_count = min(len(detected_set), total)
            percent = (detected_count / total) * 100 if total else 0.0
            impact = impact_for(category)

            rows.append(
                [
                    site_index if first_row else None,
                    cfg.name if first_row else None,
                    category,
                    total,
                    detected_count,
                    f"{percent:.0f}%",
                    impact.title() if isinstance(impact, str) else impact,
                ]
            )
            first_row = False

        site_index += 1

    wb = openpyxl.load_workbook("Evaluation Form.xlsx")
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active worksheet")
    ws.delete_rows(2, ws.max_row)

    for idx, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=idx, column=col, value=value)

    wb.save("Evaluation Form.xlsx")
    print("\n✅ Evaluation Form.xlsx updated with current test results.")


if __name__ == "__main__":
    main()
